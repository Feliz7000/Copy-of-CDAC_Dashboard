"""Admin bulk user upload."""
from pathlib import Path

from django.contrib.auth import get_user_model
from django.http import FileResponse
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from config.permissions import IsAdmin

User = get_user_model()


class UserBulkUploadViewSet(viewsets.ViewSet):
    """Admin-only bulk user creation from Excel."""

    parser_classes = (MultiPartParser, FormParser)

    def get_permissions(self):
        return [IsAdmin()]

    @action(detail=False, methods=['post'])
    def users_upload(self, request):
        if 'file' not in request.FILES:
            return Response({'detail': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']

        try:
            wb = load_workbook(file)
            ws = wb.active

            created = []
            failed = []
            errors = []

            for row_num in range(15, ws.max_row + 1):
                row = list(ws.iter_rows(
                    min_row=row_num,
                    max_row=row_num,
                    values_only=True,
                ))[0]

                if not row[0]:
                    continue

                try:
                    username = str(row[0]).strip()
                    email = str(row[1]).strip()
                    first_name = str(row[2]).strip()
                    last_name = str(row[3]).strip()
                    role = str(row[4]).strip().lower()
                    prn = row[5]
                    hod_courses = row[6]
                    password = str(row[7]).strip()

                    if not username or not email or not password:
                        raise ValueError('Username, email, and password are required')

                    if role not in ['admin', 'hod', 'faculty', 'student']:
                        raise ValueError(f'Invalid role: {role}')

                    if role == 'student' and not prn:
                        raise ValueError('PRN required for student role')

                    if User.objects.filter(username=username).exists():
                        raise ValueError(f"Username '{username}' already exists")

                    if User.objects.filter(email=email).exists():
                        raise ValueError(f"Email '{email}' already exists")

                    if role == 'student' and User.objects.filter(prn=prn).exists():
                        raise ValueError(f"PRN '{prn}' already exists")

                    User.objects.create_user(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        password=password,
                        role=role,
                        prn=prn if role == 'student' else None,
                        hod_courses=hod_courses if role == 'hod' else None,
                    )

                    created.append({
                        'row': row_num,
                        'username': username,
                        'email': email,
                        'role': role,
                    })

                except Exception as e:
                    failed.append({'row': row_num, 'error': str(e)})
                    errors.append(f'Row {row_num}: {e}')

            return Response({
                'summary': {
                    'total_processed': len(created) + len(failed),
                    'created': len(created),
                    'failed': len(failed),
                },
                'created': created[:50],
                'failed': failed[:50],
                'errors': errors[:20],
            }, status=status.HTTP_200_OK if created else status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {'detail': f'File processing error: {e}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

    @action(detail=False, methods=['get'])
    def users_template(self, request):
        file_path = Path(__file__).resolve().parent.parent.parent / 'Users_BulkUpload_Template.xlsx'
        if not file_path.exists():
            return Response({'detail': 'Template file not found'}, status=status.HTTP_404_NOT_FOUND)
        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename='Users_BulkUpload_Template.xlsx',
        )
